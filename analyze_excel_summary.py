#!/usr/bin/env python3
"""
Excel 파일 요약 분석 스크립트
워크시트별 구조와 주요 수식을 요약하여 보고합니다.
"""

import openpyxl
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET
import os

def check_vba_macros(file_path):
    """Excel 파일에서 VBA 매크로 확인"""
    print("\n=== VBA 매크로 검사 ===")
    try:
        # Excel 파일을 zip으로 열어서 vbaProject.bin 파일 확인
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            file_list = zip_file.namelist()
            vba_files = [f for f in file_list if 'vba' in f.lower() or 'macro' in f.lower()]
            
            if vba_files:
                print(f"VBA 관련 파일 발견:")
                for vba_file in vba_files:
                    print(f"  - {vba_file}")
            else:
                print("VBA 매크로가 발견되지 않았습니다.")
                
    except Exception as e:
        print(f"VBA 매크로 검사 중 오류: {e}")

def analyze_data_validation(ws):
    """데이터 유효성 검사 규칙 분석"""
    validations = []
    try:
        if hasattr(ws, 'data_validations') and ws.data_validations:
            for dv in ws.data_validations.dataValidation:
                validation = {
                    'type': dv.type,
                    'formula1': dv.formula1,
                    'formula2': dv.formula2,
                    'ranges': [str(r) for r in dv.ranges]
                }
                validations.append(validation)
    except Exception as e:
        print(f"  데이터 유효성 검사 분석 오류: {e}")
    return validations

def get_formula_summary(ws):
    """워크시트의 주요 수식 패턴 분석"""
    formulas = []
    vlookup_formulas = []
    sum_formulas = []
    reference_formulas = []
    calculation_formulas = []
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                cell_ref = f"{cell.column_letter}{cell.row}"
                
                formulas.append({'cell': cell_ref, 'formula': formula})
                
                # 수식 분류
                if 'VLOOKUP' in formula.upper():
                    vlookup_formulas.append({'cell': cell_ref, 'formula': formula})
                elif 'SUM(' in formula.upper():
                    sum_formulas.append({'cell': cell_ref, 'formula': formula})
                elif '!' in formula:  # 다른 시트 참조
                    reference_formulas.append({'cell': cell_ref, 'formula': formula})
                elif any(op in formula for op in ['+', '-', '*', '/', '%']):
                    calculation_formulas.append({'cell': cell_ref, 'formula': formula})
    
    return {
        'total_formulas': len(formulas),
        'all_formulas': formulas,
        'vlookup_formulas': vlookup_formulas,
        'sum_formulas': sum_formulas,
        'reference_formulas': reference_formulas,
        'calculation_formulas': calculation_formulas
    }

def analyze_worksheet_summary(ws, sheet_name):
    """워크시트 요약 분석"""
    print(f"\n{'='*60}")
    print(f"워크시트: {sheet_name}")
    print(f"{'='*60}")
    
    # 기본 정보
    min_row, max_row = ws.min_row, ws.max_row
    min_col, max_col = ws.min_column, ws.max_column
    used_range = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"
    
    print(f"사용 범위: {used_range}")
    print(f"행 수: {max_row - min_row + 1}")
    print(f"열 수: {max_col - min_col + 1}")
    
    # 병합된 셀
    merged_ranges = list(ws.merged_cells.ranges)
    if merged_ranges:
        print(f"\n병합된 셀 ({len(merged_ranges)}개):")
        for i, merged_range in enumerate(merged_ranges[:10]):  # 처음 10개만 표시
            print(f"  {i+1}. {merged_range}")
        if len(merged_ranges) > 10:
            print(f"  ... 및 {len(merged_ranges) - 10}개 더")
    
    # 수식 분석
    formula_analysis = get_formula_summary(ws)
    print(f"\n수식 분석:")
    print(f"  총 수식 개수: {formula_analysis['total_formulas']}")
    print(f"  VLOOKUP 수식: {len(formula_analysis['vlookup_formulas'])}개")
    print(f"  SUM 수식: {len(formula_analysis['sum_formulas'])}개")
    print(f"  시트간 참조: {len(formula_analysis['reference_formulas'])}개")
    print(f"  계산 수식: {len(formula_analysis['calculation_formulas'])}개")
    
    # 주요 수식 샘플 표시
    if formula_analysis['vlookup_formulas']:
        print(f"\n주요 VLOOKUP 수식:")
        for i, f in enumerate(formula_analysis['vlookup_formulas'][:5]):
            print(f"  {f['cell']}: {f['formula']}")
        if len(formula_analysis['vlookup_formulas']) > 5:
            print(f"  ... 및 {len(formula_analysis['vlookup_formulas']) - 5}개 더")
    
    if formula_analysis['reference_formulas']:
        print(f"\n주요 시트간 참조 수식:")
        for i, f in enumerate(formula_analysis['reference_formulas'][:5]):
            print(f"  {f['cell']}: {f['formula']}")
        if len(formula_analysis['reference_formulas']) > 5:
            print(f"  ... 및 {len(formula_analysis['reference_formulas']) - 5}개 더")
    
    # 데이터 유효성 검사
    validations = analyze_data_validation(ws)
    if validations:
        print(f"\n데이터 유효성 검사 ({len(validations)}개):")
        for i, validation in enumerate(validations):
            print(f"  {i+1}. 타입: {validation['type']}, 범위: {validation['ranges']}")
            if validation['formula1']:
                print(f"     수식1: {validation['formula1']}")
    
    # 주요 데이터 영역 식별
    print(f"\n주요 데이터 영역:")
    
    # 헤더 행 찾기 (첫 10행에서)
    headers = []
    for row in range(1, min(11, max_row + 1)):
        row_data = []
        for col in range(min_col, min_col + 10):  # 처음 10열만
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                row_data.append(f"{cell.coordinate}: {cell.value}")
        if row_data:
            headers.extend(row_data)
    
    if headers:
        print("  주요 헤더/라벨:")
        for header in headers[:10]:  # 처음 10개만
            print(f"    {header}")
        if len(headers) > 10:
            print(f"    ... 및 {len(headers) - 10}개 더")
    
    return formula_analysis

def main():
    file_path = "/Users/sung/user/workspace/GRK/GRK_workspace/2025_CF_management.xlsx"
    
    print("=== GRK Partners 2025 Cash Flow Management Excel 파일 상세 분석 ===")
    print(f"파일: {file_path}")
    
    # VBA 매크로 검사
    check_vba_macros(file_path)
    
    try:
        # Excel 파일 로드
        wb = load_workbook(file_path, data_only=False)
        
        print(f"\n=== 워크북 전체 정보 ===")
        print(f"총 워크시트 개수: {len(wb.worksheets)}")
        print(f"워크시트 목록:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        print(f"활성 시트: {wb.active.title}")
        
        # 각 워크시트 분석
        all_formulas = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formula_analysis = analyze_worksheet_summary(ws, sheet_name)
            all_formulas[sheet_name] = formula_analysis
        
        # 전체 요약
        print(f"\n{'='*60}")
        print("전체 요약")
        print(f"{'='*60}")
        
        total_formulas = sum(analysis['total_formulas'] for analysis in all_formulas.values())
        total_vlookups = sum(len(analysis['vlookup_formulas']) for analysis in all_formulas.values())
        total_references = sum(len(analysis['reference_formulas']) for analysis in all_formulas.values())
        
        print(f"전체 수식 개수: {total_formulas}")
        print(f"전체 VLOOKUP 개수: {total_vlookups}")
        print(f"전체 시트간 참조: {total_references}")
        
        # 주요 계산 로직 분석
        print(f"\n주요 계산 로직:")
        print("1. Cash Flow Management 시트:")
        print("   - 기말현금 = 기초현금 + 수입 - 지출")
        print("   - VLOOKUP으로 HR unit cost 시트에서 인력비 조회")
        print("   - 월별 비용을 합산하여 연간 총액 계산")
        
        print("2. Research CF Details 시트:")
        print("   - 프로젝트별 매출과 비용 계산")
        print("   - 운영경비 10% 자동 계산")
        print("   - 월별 손익 계산")
        
        print("3. HR unit cost 시트:")
        print("   - 직원별 월급, 상여금, 성과급 계산")
        print("   - 4대보험 및 복지비용 포함")
        
        print("4. Monthly Expense 시트:")
        print("   - 월간 고정비용 관리")
        print("   - 환율 및 직원수 기반 계산")
        
        print(f"\n분석 완료!")
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")

if __name__ == "__main__":
    main()