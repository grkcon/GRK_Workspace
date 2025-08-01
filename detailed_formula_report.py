#!/usr/bin/env python3
"""
주요 수식들의 상세 분석 리포트
"""

import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

def analyze_key_formulas(file_path):
    """주요 수식들을 카테고리별로 분석"""
    wb = load_workbook(file_path, data_only=False)
    
    print("=== 주요 수식 상세 분석 ===\n")
    
    # 01.Cash Flow Management 시트 분석
    print("1. Cash Flow Management 시트 - 핵심 계산 로직")
    print("-" * 50)
    ws = wb['01.Cash Flow Management']
    
    # 기말현금 계산 수식
    print("A. 기말현금 계산 (월별):")
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        cell = ws[f'{col}6']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {col}6: {cell.value}")
    
    # 기초현금 연결 수식
    print("\nB. 기초현금 연결 (전월 기말현금):")
    for col in ['G', 'H', 'I', 'J', 'K']:
        cell = ws[f'{col}7']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {col}7: {cell.value}")
    
    # VLOOKUP 수식 샘플
    print("\nC. 인력비 조회 VLOOKUP 수식 (HR unit cost 시트 참조):")
    vlookup_samples = ['J10', 'K10', 'L10']
    for cell_ref in vlookup_samples:
        cell = ws[cell_ref]
        if cell.value:
            print(f"   {cell_ref}: {cell.value}")
    
    # 지출 합계 계산
    print("\nD. 지출 합계 계산:")
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P']:
        cell = ws[f'{col}8']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {col}8: {cell.value}")
    
    # 시트간 참조 수식
    print("\nE. 다른 시트 참조 수식들:")
    reference_cells = ['K28', 'J32', 'J33', 'E37', 'F37', 'G37']
    for cell_ref in reference_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and '!' in cell.value:
            print(f"   {cell_ref}: {cell.value}")
    
    # Research CF Details 시트
    print(f"\n2. Research CF Details 시트 - 프로젝트 손익 계산")
    print("-" * 50)
    ws = wb['01.CF _Research CF Details']
    
    print("A. 총 Cash Flow 합계:")
    for col in ['G', 'H', 'I', 'J', 'K']:
        cell = ws[f'{col}3']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {col}3: {cell.value}")
    
    print("\nB. 운영경비 10% 계산:")
    opex_cells = ['L8', 'M8', 'N8', 'O8', 'P8']
    for cell_ref in opex_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    print("\nC. 손익 계산 (Revenue - COGS - Direct Opex):")
    income_cells = ['L9', 'M9', 'N9', 'O9', 'P9']
    for cell_ref in income_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    # HR unit cost 시트
    print(f"\n3. HR unit cost 시트 - 인력비 계산")
    print("-" * 50)
    ws = wb['03.HR unit cost']
    
    # 주요 계산 수식 찾기
    print("A. 주요 인력비 계산 수식:")
    key_cells = ['H5', 'I5', 'J5', 'K5', 'P5', 'R5', 'S5']
    for cell_ref in key_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    # Monthly Expense 시트
    print(f"\n4. Monthly Expense 시트 - 월간 비용 관리")
    print("-" * 50)
    ws = wb['02.Monthly Expense']
    
    print("A. 전체 비용 구조:")
    expense_cells = ['E5', 'E7', 'E8', 'E9']
    for cell_ref in expense_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    print("\nB. 환율 기반 계산:")
    fx_cells = ['E21', 'E22', 'E23', 'E24', 'E26']
    for cell_ref in fx_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    print("\nC. 비율 계산:")
    ratio_cells = ['G7', 'G8', 'G9', 'G10']
    for cell_ref in ratio_cells:
        cell = ws[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"   {cell_ref}: {cell.value}")
    
    # 프로젝트별 시트들
    project_sheets = ['01.SCL LIS시스템 ISP', '02.SCL HIS시스템 PMO', '03.휴니버스PMI', '99.프로젝트 PPE']
    
    print(f"\n5. 프로젝트별 시트들 - 공통 계산 구조")
    print("-" * 50)
    
    for sheet_name in project_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{sheet_name}:")
            
            # ECM 계산
            ecm_cell = ws['H5']
            if ecm_cell.value and isinstance(ecm_cell.value, str) and ecm_cell.value.startswith('='):
                print(f"   ECM 계산 (H5): {ecm_cell.value}")
            
            # VLOOKUP 인력비 계산 샘플
            vlookup_cell = ws['F18']
            if vlookup_cell.value and isinstance(vlookup_cell.value, str) and 'VLOOKUP' in vlookup_cell.value:
                print(f"   인력비 VLOOKUP (F18): {vlookup_cell.value}")
    
    print(f"\n=== 분석 완료 ===")

def main():
    file_path = "/Users/sung/user/workspace/GRK/GRK_workspace/2025_CF_management.xlsx"
    analyze_key_formulas(file_path)

if __name__ == "__main__":
    main()