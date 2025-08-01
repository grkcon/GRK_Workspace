#!/usr/bin/env python3
"""
Excel 파일의 고급 기능 분석 (차트, 피벗테이블, 데이터 유효성 검사 등)
"""

import openpyxl
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET

def analyze_charts_and_pivots(file_path):
    """차트와 피벗테이블 분석"""
    print("=== 차트 및 피벗테이블 분석 ===\n")
    
    try:
        # ZIP 파일로 열어서 내부 구조 확인
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            file_list = zip_file.namelist()
            
            # 차트 관련 파일 찾기
            chart_files = [f for f in file_list if 'chart' in f.lower()]
            if chart_files:
                print("발견된 차트 파일들:")
                for chart_file in chart_files:
                    print(f"  - {chart_file}")
            else:
                print("차트가 발견되지 않았습니다.")
            
            # 피벗테이블 관련 파일 찾기
            pivot_files = [f for f in file_list if 'pivot' in f.lower()]
            if pivot_files:
                print(f"\n발견된 피벗테이블 파일들:")
                for pivot_file in pivot_files:
                    print(f"  - {pivot_file}")
            else:
                print("\n피벗테이블이 발견되지 않았습니다.")
            
            # 드로잉 관련 파일 찾기
            drawing_files = [f for f in file_list if 'drawing' in f.lower()]
            if drawing_files:
                print(f"\n발견된 드로잉 파일들:")
                for drawing_file in drawing_files:
                    print(f"  - {drawing_file}")
            
    except Exception as e:
        print(f"ZIP 파일 분석 중 오류: {e}")

def analyze_data_validation_detailed(file_path):
    """데이터 유효성 검사 상세 분석"""
    print("\n=== 데이터 유효성 검사 상세 분석 ===\n")
    
    wb = load_workbook(file_path, data_only=False)
    
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_validations = []
        
        try:
            if hasattr(ws, 'data_validations') and ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    validation_info = {
                        'type': dv.type,
                        'formula1': dv.formula1,
                        'formula2': dv.formula2,
                        'ranges': [str(r) for r in dv.ranges],
                        'allow_blank': dv.allowBlank,
                        'show_dropdown': dv.showDropDown,
                        'error_title': dv.errorTitle,
                        'error': dv.error,
                        'prompt_title': dv.promptTitle,
                        'prompt': dv.prompt
                    }
                    sheet_validations.append(validation_info)
                    total_validations += 1
                
                if sheet_validations:
                    print(f"시트 '{sheet_name}' - {len(sheet_validations)}개의 데이터 유효성 검사:")
                    for i, validation in enumerate(sheet_validations, 1):
                        print(f"  {i}. 타입: {validation['type']}")
                        print(f"     범위: {', '.join(validation['ranges'])}")
                        if validation['formula1']:
                            print(f"     수식1: {validation['formula1']}")
                        if validation['formula2']:
                            print(f"     수식2: {validation['formula2']}")
                        if validation['error_title']:
                            print(f"     오류 제목: {validation['error_title']}")
                        if validation['error']:
                            print(f"     오류 메시지: {validation['error']}")
                        print()
        
        except Exception as e:
            print(f"  시트 '{sheet_name}' 데이터 유효성 검사 분석 오류: {e}")
    
    if total_validations == 0:
        print("데이터 유효성 검사 규칙이 발견되지 않았습니다.")
    else:
        print(f"총 {total_validations}개의 데이터 유효성 검사 규칙이 발견되었습니다.")

def analyze_conditional_formatting(file_path):
    """조건부 서식 분석"""
    print("\n=== 조건부 서식 분석 ===\n")
    
    wb = load_workbook(file_path, data_only=False)
    
    total_cf = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        try:
            if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
                cf_rules = list(ws.conditional_formatting)
                if cf_rules:
                    print(f"시트 '{sheet_name}' - {len(cf_rules)}개의 조건부 서식:")
                    for i, cf in enumerate(cf_rules, 1):
                        print(f"  {i}. 범위: {[str(r) for r in cf.cells]}")
                        print(f"     규칙 수: {len(cf.cfRule)}")
                        for j, rule in enumerate(cf.cfRule, 1):
                            print(f"       규칙 {j}: 타입={rule.type}")
                            if hasattr(rule, 'formula') and rule.formula:
                                print(f"                수식={rule.formula}")
                    total_cf += len(cf_rules)
                    print()
        
        except Exception as e:
            print(f"  시트 '{sheet_name}' 조건부 서식 분석 오류: {e}")
    
    if total_cf == 0:
        print("조건부 서식이 발견되지 않았습니다.")
    else:
        print(f"총 {total_cf}개의 조건부 서식이 발견되었습니다.")

def analyze_named_ranges(file_path):
    """명명된 범위 분석"""
    print("\n=== 명명된 범위 분석 ===\n")
    
    wb = load_workbook(file_path, data_only=False)
    
    try:
        if hasattr(wb, 'defined_names') and wb.defined_names:
            named_ranges = list(wb.defined_names)
            if named_ranges:
                print(f"발견된 명명된 범위 ({len(named_ranges)}개):")
                for named_range in named_ranges:
                    print(f"  이름: {named_range.name}")
                    print(f"  범위: {named_range.attr_text}")
                    if named_range.localSheetId is not None:
                        sheet_names = wb.sheetnames
                        if named_range.localSheetId < len(sheet_names):
                            print(f"  시트: {sheet_names[named_range.localSheetId]}")
                    print()
            else:
                print("명명된 범위가 발견되지 않았습니다.")
        else:
            print("명명된 범위가 발견되지 않았습니다.")
    
    except Exception as e:
        print(f"명명된 범위 분석 중 오류: {e}")

def analyze_workbook_properties(file_path):
    """워크북 속성 분석"""
    print("\n=== 워크북 속성 분석 ===\n")
    
    wb = load_workbook(file_path, data_only=False)
    
    try:
        # 워크북 속성
        properties = wb.properties
        if properties:
            print("워크북 속성:")
            print(f"  제목: {properties.title}")
            print(f"  작성자: {properties.creator}")
            print(f"  마지막 수정자: {properties.lastModifiedBy}")
            print(f"  생성일: {properties.created}")
            print(f"  수정일: {properties.modified}")
            print(f"  주제: {properties.subject}")
            print(f"  설명: {properties.description}")
            print(f"  키워드: {properties.keywords}")
            print(f"  카테고리: {properties.category}")
        
        # 보안 설정
        if hasattr(wb, 'security') and wb.security:
            print(f"\n보안 설정:")
            print(f"  워크북 보호: {wb.security.workbookPassword is not None}")
            print(f"  구조 보호: {wb.security.lockStructure}")
            print(f"  창 보호: {wb.security.lockWindows}")
    
    except Exception as e:
        print(f"워크북 속성 분석 중 오류: {e}")

def main():
    file_path = "/Users/sung/user/workspace/GRK/GRK_workspace/2025_CF_management.xlsx"
    
    # 차트 및 피벗테이블 분석
    analyze_charts_and_pivots(file_path)
    
    # 데이터 유효성 검사 분석
    analyze_data_validation_detailed(file_path)
    
    # 조건부 서식 분석
    analyze_conditional_formatting(file_path)
    
    # 명명된 범위 분석
    analyze_named_ranges(file_path)
    
    # 워크북 속성 분석
    analyze_workbook_properties(file_path)
    
    print("\n=== 고급 기능 분석 완료 ===")

if __name__ == "__main__":
    main()