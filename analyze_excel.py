#!/usr/bin/env python3
"""
Excel 파일 상세 분석 스크립트
모든 워크시트, 수식, 데이터 구조, 계산 로직을 추출합니다.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import sys

def analyze_excel_file(file_path):
    """Excel 파일을 상세히 분석합니다."""
    try:
        # Excel 파일 로드 (수식 유지)
        wb = load_workbook(file_path, data_only=False)
        
        analysis = {
            "file_path": file_path,
            "worksheets": [],
            "workbook_info": {
                "sheet_names": wb.sheetnames,
                "active_sheet": wb.active.title if wb.active else None,
                "total_sheets": len(wb.worksheets)
            }
        }
        
        print(f"=== Excel 파일 분석: {file_path} ===\n")
        print(f"총 워크시트 개수: {len(wb.worksheets)}")
        print(f"워크시트 목록: {wb.sheetnames}")
        print(f"활성 시트: {wb.active.title if wb.active else 'None'}\n")
        
        # 각 워크시트 분석
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_analysis = analyze_worksheet(ws, sheet_name)
            analysis["worksheets"].append(sheet_analysis)
            
        return analysis
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return None

def analyze_worksheet(ws, sheet_name):
    """개별 워크시트를 분석합니다."""
    print(f"\n{'='*50}")
    print(f"워크시트: {sheet_name}")
    print(f"{'='*50}")
    
    # 사용된 셀 범위 확인
    min_row, max_row = ws.min_row, ws.max_row
    min_col, max_col = ws.min_column, ws.max_column
    
    print(f"사용된 범위: {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}")
    print(f"행 범위: {min_row} ~ {max_row}")
    print(f"열 범위: {get_column_letter(min_col)} ~ {get_column_letter(max_col)}")
    
    sheet_data = {
        "name": sheet_name,
        "dimensions": {
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
            "used_range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        },
        "formulas": [],
        "data": [],
        "merged_cells": [],
        "data_validation": [],
        "conditional_formatting": [],
        "charts": [],
        "pivot_tables": []
    }
    
    # 병합된 셀 정보
    if ws.merged_cells.ranges:
        print(f"\n병합된 셀:")
        for merged_range in ws.merged_cells.ranges:
            print(f"  - {merged_range}")
            sheet_data["merged_cells"].append(str(merged_range))
    
    # 모든 셀 데이터 및 수식 분석
    print(f"\n셀 데이터 및 수식 분석:")
    formulas_found = []
    
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell_ref = f"{get_column_letter(col)}{row}"
            
            cell_info = {
                "address": cell_ref,
                "value": cell.value,
                "data_type": str(type(cell.value).__name__),
                "formula": None,
                "number_format": cell.number_format,
                "font": {
                    "name": cell.font.name,
                    "size": cell.font.size,
                    "bold": cell.font.bold,
                    "italic": cell.font.italic
                },
                "fill": str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb != '00000000' else None,
                "alignment": {
                    "horizontal": cell.alignment.horizontal,
                    "vertical": cell.alignment.vertical
                }
            }
            
            # 수식이 있는 셀 확인
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                cell_info["formula"] = cell.value
                formulas_found.append({
                    "cell": cell_ref,
                    "formula": cell.value
                })
                print(f"  {cell_ref}: {cell.value}")
            elif cell.value is not None:
                # 값이 있는 셀만 출력 (너무 많은 출력 방지)
                if row <= min_row + 10:  # 처음 몇 행만 출력
                    print(f"  {cell_ref}: {cell.value} ({type(cell.value).__name__})")
            
            row_data.append(cell_info)
        sheet_data["data"].append(row_data)
    
    # 수식 요약
    if formulas_found:
        print(f"\n발견된 수식 총 {len(formulas_found)}개:")
        sheet_data["formulas"] = formulas_found
        for formula_info in formulas_found:
            print(f"  {formula_info['cell']}: {formula_info['formula']}")
    else:
        print(f"\n수식이 발견되지 않았습니다.")
    
    # 데이터 유효성 검사 규칙
    try:
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"\n데이터 유효성 검사:")
            for dv in ws.data_validations.dataValidation:
                validation_info = {
                    "type": dv.type,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "ranges": [str(r) for r in dv.ranges]
                }
                sheet_data["data_validation"].append(validation_info)
                print(f"  타입: {dv.type}, 범위: {[str(r) for r in dv.ranges]}")
                if dv.formula1:
                    print(f"    수식1: {dv.formula1}")
                if dv.formula2:
                    print(f"    수식2: {dv.formula2}")
    except Exception as e:
        print(f"  데이터 유효성 검사 분석 오류: {e}")
    
    # 조건부 서식
    try:
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            print(f"\n조건부 서식:")
            for cf in ws.conditional_formatting:
                cf_info = {
                    "type": str(type(cf)),
                    "ranges": [str(r) for r in cf.cells]
                }
                sheet_data["conditional_formatting"].append(cf_info)
                print(f"  {cf_info}")
    except Exception as e:
        print(f"  조건부 서식 분석 오류: {e}")
    
    # 차트 정보
    try:
        if hasattr(ws, '_charts') and ws._charts:
            print(f"\n차트:")
            for chart in ws._charts:
                chart_info = {
                    "type": str(type(chart)),
                    "title": getattr(chart, 'title', 'Unknown')
                }
                sheet_data["charts"].append(chart_info)
                print(f"  {chart_info}")
    except Exception as e:
        print(f"  차트 분석 오류: {e}")
    
    return sheet_data

def main():
    file_path = "/Users/sung/user/workspace/GRK/GRK_workspace/2025_CF_management.xlsx"
    
    print("Excel 파일 상세 분석을 시작합니다...\n")
    
    # Excel 파일 분석
    analysis = analyze_excel_file(file_path)
    
    if analysis:
        # JSON 형태로도 저장
        output_file = "/Users/sung/user/workspace/GRK/GRK_workspace/excel_analysis.json"
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(analysis, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n분석 결과가 {output_file}에 저장되었습니다.")
        except Exception as e:
            print(f"JSON 저장 오류: {e}")
    
    print("\n분석 완료!")

if __name__ == "__main__":
    main()